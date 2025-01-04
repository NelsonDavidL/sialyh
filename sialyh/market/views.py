import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db import connection
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from .forms import *
from .models import birthday_list

# Create your views here.

@login_required
def view_market(request):
    return render(request, "market/market.html")

@login_required
def upload_file(request):
    if request.method == 'POST':
        form = form_upload_file(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES ['file']
            month_form = form.cleaned_data['month']
            try:
                df = pd.read_csv(file, delimiter=';')

                list_birthday = []
                for index, row in df.iterrows():
                    convert_user =  row['TIENDA'].replace("-", "")
                    out_phone = str(row['TEL'])
                    convert_obs = row['OBSERVACION'] if pd.notna(row['OBSERVACION']) else ""
                    if not out_phone.isdigit():
                        continue
                    list_birthday.append(birthday_list(
                        user = convert_user,
                        day = row['DIA'],
                        line = row['LINEA'],
                        client = row['CLIENTE'],
                        phone = out_phone,
                        observation = convert_obs,
                        month = month_form,
                    ))
            
                birthday_list.objects.bulk_create(list_birthday)
                messages.success(request, "Datos importados correctamente")
                return redirect ('upload_list')
            except Exception as e:
                print(f"Este es el error: {e}")
                messages.error(request, f"Error al importar datos: {e}")
    else:
        form = form_upload_file()
    return render (request, 'market/upload_list.html', {'form':form})

@login_required
def truncate_list(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("TRUNCATE TABLE public.birthday_list RESTART IDENTITY;")
        messages.success(request, 'Base limpiada correctamente')
    except Exception as e:
        messages.error(request, f"Error al limpiar la base {e}")
    return redirect('upload_list')

@login_required
def list_birthday(request):

    form = birthday_list.objects.filter(user=request.user)
    first_record = birthday_list.objects.order_by('id').first()
    first_month = first_record.month if first_record else None
    return render(request, 'market/list_birthday.html', {'form':form, 'first_month':first_month})

@login_required
def insert_observation(request, birth_id):
    observation = request.POST.get('observation')
    birthday_list.objects.filter(id=birth_id).update(observation=observation)
    return redirect('list_birthday')

@login_required
def export_list_birthday(request):

    # Crear el libro de trabajo
    workbook = Workbook()
    # Seleccionar la hoja activa
    sheet = workbook.active
    sheet.title = 'Birthday List'

    # Escribir los encabezados
    headers = ['Tienda', 'Día', 'Línea', 'Cliente', 'Teléfono', 'Observación']
    sheet.append(headers)

    # Obtener todos los registros de birthday_list
    queryset = birthday_list.objects.all().order_by('id')

    # Escribir los datos
    for birthday in queryset:
        row = [
            birthday.user,
            birthday.day,
            birthday.line,
            birthday.client,
            birthday.phone,
            birthday.observation
        ]
        sheet.append(row)


    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Obtener la letra de la columna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Crear una respuesta HTTP con el tipo de contenido de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=birthday_list.xlsx'

    # Guardar el archivo en la respuesta
    workbook.save(response)

    return response