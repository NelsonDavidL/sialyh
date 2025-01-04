import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from .forms import *
from django.contrib.auth.decorators import login_required
from .models import *
from django.contrib import messages
from django.db import connection
from django.views.generic import UpdateView
from django.urls import reverse_lazy
from django.utils import timezone
from django.http import HttpResponse, JsonResponse

from datetime import datetime
from sialyhapp.functions import *


# Create your views here.

# ACA VAN LAS VISTAS PARA LOS MODULOS DE INGRESOS Y LISTADOS

# Esta es la vista de la seccion de facturacion en el Home
@login_required
def billingView(request):
    return render(request, "billing/billing.html")


# Esta es la vista que renderiza las opciones para los ingresos
@login_required
def incomeView(request):
    return render(request, "billing/income.html")


# Esta es la vista que renderiza las opciones para los listados
@login_required
def listingsView(request):
    return render(request, "billing/listings.html")

@login_required
def bidBonsView(request):
    return render(request, "billing/bid_bonds.html")


# ACA VAN LAS VISTAS DE LOS FORMULARIOS DE INGRESO DE DOCUMENTOS

# Esta es la vista del formulario de ingreso de facturas manuales
@login_required
def manual_invoicesView(request):
    if request.method == 'POST':
        form = form_manual_billing(request.POST, request.FILES)
        form_list = manInvoice.objects.filter(user=request.user)
        if form.is_valid():
            man_invoice = form.save(commit=False, user=request.user)
            man_invoice.user = request.user
            man_invoice.save()
            messages.success(request, 'Información enviada correctamente.')
            return redirect('manual_invoices')
    else:
        form = form_manual_billing()
        form_list = manInvoice.objects.filter(user=request.user)
    return render(request, 'manual_invoices/manual_invoices.html', {
        'form': form,
        'form_list':form_list
        })


# Esta es la vista del formulario de ingreso de bonos de separado
@login_required
def sep_bonusesView(request):
    if request.method == 'POST':
        formsb = form_separated_bonuses(request.POST, request.FILES)
        formsb_list = sepBonuses.objects.filter(user=request.user)
        if formsb.is_valid():
            separated_bond = formsb.save(commit=False, user=request.user)
            separated_bond.user = request.user
            separated_bond.save()
            messages.success(request, 'Información enviada correctamente.')
            return redirect('sepBonuses')
    else:
        formsb = form_separated_bonuses()
        formsb_list = sepBonuses.objects.filter(user=request.user)
    return render(request, 'separate_bonuses/separate_bonuses.html', {
        'formsb': formsb,
        'formsb_list':formsb_list
        })


# Esta es la vista del formulario de ingreso de descuentos por nomina
@login_required
def pay_discountsView(request):
    if request.method == 'POST':
        formpp = form_payroll_discount(request.POST, request.FILES)
        formpp_list = payDiscounts.objects.filter(user=request.user)

        if formpp.is_valid():
            payroll_discount = formpp.save(commit=False, user=request.user)
            payroll_discount.user = request.user
            payroll_discount.save()
            messages.success(request, 'Información enviada correctamente.')
            return redirect('payDiscounts')
    else:
        formpp = form_payroll_discount()
        formpp_list = payDiscounts.objects.filter(user=request.user)
    return render(request, 'payroll_discounts/payroll_discounts.html', {
        'formpp': formpp,
        'formpp_list':formpp_list
        })


#Esta es la vista del formulario de ingreso de manejo de datos
@login_required
def data_managementView(request):
    # Lógica para el formulario de envío de datos
    if request.method == 'POST':
        formdm = form_data_management(request.POST, request.FILES)
        formdm_list = dataManagement.objects.filter(user=request.user)
        if formdm.is_valid():
            data_management = formdm.save(commit=False, user=request.user)
            data_management.user = request.user
            data_management.save()
            messages.success(request, 'Información enviada correctamente.')
            return redirect('dataManagement')
    else:
        formdm = form_data_management()
        formdm_list = dataManagement.objects.filter(user=request.user)
    return render(request, 'data_management/data_management.html', {
        'formdm': formdm,
        'formdm_list':formdm_list
        })


# ACA VAN LAS VISTAS DE LOS LISTADOS DE LOS DOCUMENTOS INGRESADOS

# Esta es la vista del listado de facturas manuales
@login_required
def invoice_list(request):
    
    miForm = manInvoice.objects.all()
    # Aca se inicializan los formularios de filtro
    start_date_filter_form = form_start_date(request.GET)
    end_date_filter_form = form_end_date(request.GET)
    user_filter_form = form_user_filter(request.GET)
    # Aca se verifica si los formularios son válidos
    if start_date_filter_form.is_valid() and end_date_filter_form.is_valid() and user_filter_form.is_valid():
        
        start_date_invoice = start_date_filter_form.cleaned_data.get('start_date')
        end_date_invoice = end_date_filter_form.cleaned_data.get('end_date')
        user_filter = user_filter_form.cleaned_data.get('user')
        # Aca se filtra el queryset de manInvoice
        if start_date_invoice and end_date_invoice:
            miForm = miForm.filter(created__range=(start_date_invoice, end_date_invoice))
        if user_filter:
            miForm = miForm.filter(user__username=user_filter)
        # Aca se renderiza la plantilla con los objetos filtrados y los formularios de filtro
        return render(request, 'manual_invoices/invoices.html', {
            'miForm': miForm,
            'start_date_filter_form': start_date_filter_form,
            'end_date_filter_form': end_date_filter_form,
            'user_filter_form': user_filter_form,
            'start_filtered_date': start_date_invoice,
            'end_filtered_date': end_date_invoice,
            'filtered_user': user_filter,
        })
    else:
        return render(request, 'manual_invoices/invoices.html', {
            'miForm': miForm,
            'start_date_filter_form': start_date_filter_form,
            'end_date_filter_form': end_date_filter_form,
            'user_filter_form': user_filter_form,
        })


# Esta es la vista del listado de bonos de separado
@login_required
def sepBonuses_list(request):
    
    miForm = sepBonuses.objects.all()
    # Aca se inicializan los formularios de filtro
    start_date_filter_form = form_start_date(request.GET)
    end_date_filter_form = form_end_date(request.GET)
    user_filter_form = form_user_filter(request.GET)
    # Aca se verifica si los formularios son válidos
    if start_date_filter_form.is_valid() and end_date_filter_form.is_valid() and user_filter_form.is_valid():
        
        start_date_invoice = start_date_filter_form.cleaned_data.get('start_date')
        end_date_invoice = end_date_filter_form.cleaned_data.get('end_date')
        user_filter = user_filter_form.cleaned_data.get('user')
        # Aca se filtra el queryset de manInvoice
        if start_date_invoice and end_date_invoice:
            miForm = miForm.filter(created__range=(start_date_invoice, end_date_invoice))
        if user_filter:
            miForm = miForm.filter(user__username=user_filter)
        # Aca se renderiza la plantilla con los objetos filtrados y los formularios de filtro
        return render(request, 'separate_bonuses/separate_bonuses_list.html', {
            'miForm': miForm,
            'start_date_filter_form': start_date_filter_form,
            'end_date_filter_form': end_date_filter_form,
            'user_filter_form': user_filter_form,
            'start_filtered_date': start_date_invoice,
            'end_filtered_date': end_date_invoice,
            'filtered_user': user_filter,
        })
    else:
        return render(request, 'separate_bonuses/separate_bonuses_list.html', {
            'miForm': miForm,
            'start_date_filter_form': start_date_filter_form,
            'end_date_filter_form': end_date_filter_form,
            'user_filter_form': user_filter_form,
        })


# Esta es la vista del listado de los descuentos por nómina
@login_required
def payDiscounts_list(request):
    
    miForm = payDiscounts.objects.all()
    # Aca se inicializan los formularios de filtro
    start_date_filter_form = form_start_date(request.GET)
    end_date_filter_form = form_end_date(request.GET)
    user_filter_form = form_user_filter(request.GET)
    # Aca se verifica si los formularios son válidos
    if start_date_filter_form.is_valid() and end_date_filter_form.is_valid() and user_filter_form.is_valid():
        
        start_date_invoice = start_date_filter_form.cleaned_data.get('start_date')
        end_date_invoice = end_date_filter_form.cleaned_data.get('end_date')
        user_filter = user_filter_form.cleaned_data.get('user')
        # Aca se filtra el queryset de manInvoice
        if start_date_invoice and end_date_invoice:
            miForm = miForm.filter(created__range=(start_date_invoice, end_date_invoice))
        if user_filter:
            miForm = miForm.filter(user__username=user_filter)
        # Aca se renderiza la plantilla con los objetos filtrados y los formularios de filtro
        return render(request, 'payroll_discounts/payroll_discounts_list.html', {
            'miForm': miForm,
            'start_date_filter_form': start_date_filter_form,
            'end_date_filter_form': end_date_filter_form,
            'user_filter_form': user_filter_form,
            'start_filtered_date': start_date_invoice,
            'end_filtered_date': end_date_invoice,
            'filtered_user': user_filter,
        })
    else:
        return render(request, 'payroll_discounts/payroll_discounts_list.html', {
            'miForm': miForm,
            'start_date_filter_form': start_date_filter_form,
            'end_date_filter_form': end_date_filter_form,
            'user_filter_form': user_filter_form,
        })


# Esta es la vista del listado de manejo de datos personales
@login_required
def dataManagement_list(request):
    miFormdm = dataManagement.objects.all()
    # Aca se inicializan los formularios de filtro
    date_filter_formdm = date_filterFormdm(request.GET)
    user_filter_formdm = user_filterFormdm(request.GET)

    # Aca se verifica si los formularios son válidos
    if date_filter_formdm.is_valid() and user_filter_formdm.is_valid():
        date = date_filter_formdm.cleaned_data.get('date')
        user = user_filter_formdm.cleaned_data.get('user')

        # Aca se filtra el queryset de sepBonuses
        if date:
            miFormdm = miFormdm.filter(created=date)
        if user:
            miFormdm = miFormdm.filter(user__username=user)

        return render(request, 'data_management/data_management_list.html', {
            'miFormdm': miFormdm,
            'date_filter_formdm': date_filter_formdm,
            'user_filter_formdm': user_filter_formdm,
            'filtered_date': date,
            'filtered_user': user,
        })
    # Aca se renderiza la plantilla con los objetos filtrados y los formularios de filtro
    return render(request, 'data_management/data_management_list.html', {
        'miFormdm': miFormdm,
        'date_filter_formdm': date_filter_formdm,
        'user_filter_formdm': user_filter_formdm,})


# ACA VAN LAS VISTAS PARA DESHABILITAR LOS ITEMS DE LOS LISTADOS

# Esta es la vista para deshabilitar items de las facturas manuales
@login_required
def disable_item_manInvoice(request, invoice_id):
    invoice = manInvoice.objects.get(id=invoice_id)
    invoice.enabled = not invoice.enabled
    invoice.save()
    return redirect('invoices')


# Esta es la vista para deshabilitar items de los bonos de separado
@login_required
def disable_item_sepBonuses(request, bond_id):
    bond = sepBonuses.objects.get(id=bond_id)
    bond.enabled = not bond.enabled
    bond.save()
    return redirect('sepBonuseslist')


# Esta es la vista para deshabilitar items de las descuentos por nomina
@login_required
def disable_item_payDiscount(request, payd_id):
    payd = payDiscounts.objects.get(id=payd_id)
    payd.enabled = not payd.enabled
    payd.save()
    return redirect('payDiscountslist')


# ACA VAN LAS VISTAS PARA EXPORTAR LOS LISTADOS A EXCEL

# Esta es la vista que exporta el listado de facturas manuales a excel
@login_required
def export_invoices_to_excel(request):

    print("Entrando en la función export_invoices_to_excel")

    invoices = manInvoice.objects.all()

    if request.method == 'POST':

        start_date_filter_form = form_start_date(request.GET)
        end_date_filter_form = form_end_date(request.GET)
        user_filter_form = form_user_filter(request.GET)

        if start_date_filter_form.is_valid() and end_date_filter_form.is_valid() and user_filter_form.is_valid():
            start_date_invoice = start_date_filter_form.cleaned_data.get('start_date')
            end_date_invoice = end_date_filter_form.cleaned_data.get('end_date')
            user_filter = user_filter_form.cleaned_data.get('user')

        start_date_invoice = request.POST.get('start_date')
        end_date_invoice = request.POST.get('end_date')
        user_filter = request.POST.get('user')

        print("Método POST detectado")
        print(f"Fecha de factura inicial ingresada: {start_date_invoice}")
        print(f"Fecha de factura final ingresada: {end_date_invoice}")
        print(f"Filtro de usuario ingresado: {user_filter}")

        if start_date_invoice and end_date_invoice or user_filter:

            try:
                start_date = convert_date(start_date_invoice)
                end_date = convert_date(end_date_invoice)
                print(f"Fecha de factura inicial convertida a objeto date: {start_date}")
                print(f"Fecha de factura final convertida a objeto date: {end_date}")

                invoices = manInvoice.objects.filter(created__range=(start_date, end_date), enabled=True)
                print("Filtro de fecha aplicado")
            except ValueError:
                start_date = None
                end_date = None

                user = user_filter
                print("Aplicando filtro por usuario")
                invoices = manInvoice.objects.filter(user__username=user, enabled=True)
                print("Filtro de usuario aplicado")

            response = generate_excel_invoices(invoices, request)
            return response


# Esta es la vista que exporta el listado de bonos de separado a excel
@login_required
def export_sepBonuses_to_excel(request):

    sepBond = sepBonuses.objects.all()

    if request.method == 'POST':

        start_date_filter_form = form_start_date(request.GET)
        end_date_filter_form = form_end_date(request.GET)
        user_filter_form = form_user_filter(request.GET)

        if start_date_filter_form.is_valid() and end_date_filter_form.is_valid() and user_filter_form.is_valid():
            start_date_invoice = start_date_filter_form.cleaned_data.get('start_date')
            end_date_invoice = end_date_filter_form.cleaned_data.get('end_date')
            user_filter = user_filter_form.cleaned_data.get('user')

        start_date_invoice = request.POST.get('start_date')
        end_date_invoice = request.POST.get('end_date')
        user_filter = request.POST.get('user')

        print("Método POST detectado")
        print(f"Fecha de factura inicial ingresada: {start_date_invoice}")
        print(f"Fecha de factura final ingresada: {end_date_invoice}")
        print(f"Filtro de usuario ingresado: {user_filter}")

        if start_date_invoice and end_date_invoice or user_filter:

            try:
                start_date = convert_date(start_date_invoice)
                end_date = convert_date(end_date_invoice)
                print(f"Fecha de factura inicial convertida a objeto date: {start_date}")
                print(f"Fecha de factura final convertida a objeto date: {end_date}")

                sepBond = sepBonuses.objects.filter(created__range=(start_date, end_date), enabled=True)
                print("Filtro de fecha aplicado")
            except ValueError:
                start_date = None
                end_date = None

                user = user_filter
                print("Aplicando filtro por usuario")
                sepBond = sepBonuses.objects.filter(user__username=user, enabled=True)
                print("Filtro de usuario aplicado")

            response = generate_excel_bonuses(sepBond, request)
            return response


# Esta es la vista que exporta el listado de descuentos por nómina a excel
@login_required
def export_payDiscounts_to_excel(request):

    payDiscount = payDiscounts.objects.all()

    if request.method == 'POST':

        start_date_filter_form = form_start_date(request.GET)
        end_date_filter_form = form_end_date(request.GET)
        user_filter_form = form_user_filter(request.GET)

        if start_date_filter_form.is_valid() and end_date_filter_form.is_valid() and user_filter_form.is_valid():
            start_date_invoice = start_date_filter_form.cleaned_data.get('start_date')
            end_date_invoice = end_date_filter_form.cleaned_data.get('end_date')
            user_filter = user_filter_form.cleaned_data.get('user')

        start_date_invoice = request.POST.get('start_date')
        end_date_invoice = request.POST.get('end_date')
        user_filter = request.POST.get('user')

        print("Método POST detectado")
        print(f"Fecha de factura inicial ingresada: {start_date_invoice}")
        print(f"Fecha de factura final ingresada: {end_date_invoice}")
        print(f"Filtro de usuario ingresado: {user_filter}")

        if start_date_invoice and end_date_invoice or user_filter:

            try:
                start_date = convert_date(start_date_invoice)
                end_date = convert_date(end_date_invoice)
                print(f"Fecha de factura inicial convertida a objeto date: {start_date}")
                print(f"Fecha de factura final convertida a objeto date: {end_date}")

                payDiscount = payDiscounts.objects.filter(created__range=(start_date, end_date), enabled=True)
                print("Filtro de fecha aplicado")
            except ValueError:
                start_date = None
                end_date = None

                user = user_filter
                print("Aplicando filtro por usuario")
                payDiscount = payDiscounts.objects.filter(user__username=user, enabled=True)
                print("Filtro de usuario aplicado")

            response = generate_excel_payDiscounts(payDiscount, request)
            return response

# Esta es la vista que exporta el listado de manejo de datos a excel
@login_required
def export_dataManagement_to_excel(request):

    data = dataManagement.objects.all()

    if request.method == 'POST':

        date_filter_formdm = date_filterFormdm(request.GET)
        user_filter_formdm = user_filterFormdm(request.GET)

        if date_filter_formdm.is_valid() and user_filter_formdm.is_valid():
            date_data = date_filter_formdm.cleaned_data.get('date')
            user_data = user_filter_formdm.cleaned_data.get('user')

        date_data = request.POST.get('date')
        user_data = request.POST.get('user')

        print("Método POST detectado")
        print(f"Fecha de factura ingresada: {date_data}")
        print(f"Filtro de usuario ingresado: {user_data}")

        if date_data:

            try:
                date = convert_date(date_data)
                print(f"Fecha de factura convertida a objeto date: {date}")
                print(f"Aplicando filtro por fecha de factura: {date}")
                data = dataManagement.objects.filter(created=date)
                print("Filtro de fecha aplicado")
            except ValueError:
                date_data = None

                user = user_data
                print("Aplicando filtro por usuario")
                data = dataManagement.objects.filter(user__username=user)
                print("Filtro de usuario aplicado")

        response = generate_excel_dataManagement(data, request)
        return response

@login_required
def upload_bonds(request):
    if request.method == 'POST':
        form = form_upload_file(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES ['file']
            try:
                df = pd.read_csv(file, delimiter=';')

                required_fields = ['licitacion', 'cedula', 'bono', 'valor']
                if not all (col in df.columns for col in required_fields):
                    messages.error(request, "El archivo no tiene las columnas requeridas.")
                    return redirect('upload_bonds')
                
                for _,  row in df.iterrows():
                    out_value = row['valor']
                    if not str(out_value).replace('.', '', 1).isdigit():
                        out_value = None
                    out_bond = row['bono'] if pd.notnull(row['bono']) and row['bono'] != '' else None
                    bindBonds.objects.create(
                        bidding=row['licitacion'],
                        idCard=row['cedula'],
                        numBond=out_bond,
                        value=out_value,                        
                    )
                
                messages.success(request, "Los bonos se han cargado correctamente.")
                return redirect('upload_bonds')
            except Exception as e:
                messages.error(request, f"Error procesando el archivo: {e}")
        
    else:
        form = form_upload_file()

    return render (request, 'bid_bonds/upload_bonds.html', {'form':form})

@login_required
def export_list_bond(request):

    # Crear el libro de trabajo
    workbook = Workbook()
    # Seleccionar la hoja activa
    sheet = workbook.active
    sheet.title = 'Bonos de licitación'

    # Escribir los encabezados
    headers = ['Licitacion', 'Cedula', 'Bono', 'Valor', 'Estado', 'Fecha de redimido', 'Tienda']
    sheet.append(headers)

    # Obtener todos los registros de birthday_list
    queryset = bindBonds.objects.all().order_by('id')

    # Escribir los datos
    for bond in queryset:
        row = [
            bond.bidding,
            bond.idCard,
            bond.numBond,
            bond.value,
            bond.state,
            bond.redemptionDate if bond.redemptionDate else 'N/A',
            bond.user.username if bond.user else 'N/A',
        ]
        sheet.append(row)


    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Obtener la letra de la columna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Crear una respuesta HTTP con el tipo de contenido de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=bond_list.xlsx'

    # Guardar el archivo en la respuesta
    workbook.save(response)

    return response

@login_required
def bind_bond_list(request):

    id_card_filter = form_filter_idcard(request.GET)
    num_bond_filter = form_filter_num_bond(request.GET)

    bonds = []

    if id_card_filter.is_valid() or num_bond_filter.is_valid():
        
        if id_card_filter.is_valid():
            id_card = id_card_filter.cleaned_data.get('id_card')
            if id_card: 
                bonds = bindBonds.objects.filter(idCard=id_card)

        if num_bond_filter.is_valid():
            num_bond = num_bond_filter.cleaned_data.get('num_bond')
            if num_bond:  
                if bonds:
                    
                    bonds = bonds.filter(numBond=num_bond)
                else:
                    
                    bonds = bindBonds.objects.filter(numBond__icontains=num_bond)

    return render(request, 'bid_bonds/bonds_list.html', {
        'bonds': bonds,
        'id_card_filter': id_card_filter,
        'num_bond_filter': num_bond_filter,
    })


@login_required
def update_bond(request, bond_id):
    if request.method == "POST":
        # Obtiene el bono específico o lanza un 404
        bond = get_object_or_404(bindBonds, id=bond_id)

        # Recupera los valores del formulario
        bidding_bond = request.POST.get('bidding')
        id_card_bond = request.POST.get('idCard')
        num_bond = request.POST.get('numBond')
        val_bond = request.POST.get('value')
        state_bond = request.POST.get('state')

        # Log para identificar los valores que llegan

        print(f"bidding: '{bidding_bond}'")
        print(f"idCard: '{id_card_bond}'")
        print(f"numBond: '{num_bond}'")
        print(f"value: '{val_bond}'")
        print(f"state: '{state_bond}'")

        # Valida que todos los campos tengan un valor
        if not num_bond.strip() or not val_bond.strip() or not state_bond.strip():
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect('bind_bond_list')
        
        try:
            # Actualiza el registro
            bond.bidding = bidding_bond
            bond.idCard = id_card_bond
            bond.numBond = num_bond
            bond.value = val_bond
            bond.state = state_bond

            if state_bond == 'redimido':
                bond.redemptionDate = timezone.now().date()
            
            bond.user=request.user

            bond.save()

            # Mensaje de éxito
            messages.success(request, "Bono actualizado correctamente.")
        except Exception as e:
            # Manejo de errores
            messages.error(request, f"Error al actualizar el bono: {e}")

        return redirect('bind_bond_list')
    
    # Si el método no es POST
    messages.error(request, "Método no permitido.")
    return redirect('bind_bond_list')


@login_required
def bondManagement(request):
    # Obtener todos los bonos inicialmente
    form = bindBonds.objects.all()

    # Crear los filtros a partir de los parámetros GET
    id_card_filter = form_filter_idcard(request.GET)
    num_bond_filter = form_filter_num_bond(request.GET)

    # Solo aplicar los filtros si son válidos
    if id_card_filter.is_valid() or num_bond_filter.is_valid():
        # Filtro por Cédula (id_card)
        if id_card_filter.is_valid():
            id_card = id_card_filter.cleaned_data.get('id_card')
            if id_card:
                form = form.filter(idCard=id_card)
        
        # Filtro por Número de Bono (num_bond)
        if num_bond_filter.is_valid():
            num_bond = num_bond_filter.cleaned_data.get('num_bond')
            if num_bond:
                form = form.filter(numBond__icontains=num_bond)

    return render(request, 'bid_bonds/bond_management.html', {
        'form': form,
        'id_card_filter': id_card_filter,
        'num_bond_filter': num_bond_filter,
    })

@login_required
def management_update_bond(request, bond_id):
    if request.method == "POST":
        # Obtiene el bono específico o lanza un 404
        bond = get_object_or_404(bindBonds, id=bond_id)

        # Recupera los valores del formulario
        bidding_bond = request.POST.get('bidding')
        id_card_bond = request.POST.get('idCard')
        num_bond = request.POST.get('numBond')
        val_bond = request.POST.get('value')
        state_bond = request.POST.get('state')

        # Log para identificar los valores que llegan

        print(f"bidding: '{bidding_bond}'")
        print(f"idCard: '{id_card_bond}'")
        print(f"numBond: '{num_bond}'")
        print(f"value: '{val_bond}'")
        print(f"state: '{state_bond}'")

        # Valida que todos los campos tengan un valor
        if not num_bond.strip() or not val_bond.strip() or not state_bond.strip():
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect('bond_management')
        
        try:
            # Actualiza el registro
            bond.bidding = bidding_bond
            bond.idCard = id_card_bond
            bond.numBond = num_bond
            bond.value = val_bond
            bond.state = state_bond

            if state_bond == 'redimido':
                bond.redemptionDate = timezone.now().date()
            
            bond.user=request.user

            bond.save()

            # Mensaje de éxito
            messages.success(request, "Bono actualizado correctamente.")
        except Exception as e:
            # Manejo de errores
            messages.error(request, f"Error al actualizar el bono: {e}")

        return redirect('bond_management')
    
    # Si el método no es POST
    messages.error(request, "Método no permitido.")
    return redirect('bond_management')

@login_required
def add_bond(request, id_Card):
    form = bindBonds.objects.filter(idCard=id_Card)
    form_add = form_add_bond()

    if request.method == 'POST':
        form_add = form_add_bond(request.POST)  # Pasar los datos enviados

        if form_add.is_valid():
            bidding_bond = bindBonds.objects.filter(idCard=id_Card).values_list('bidding').first()
            id_card_bond = id_Card
            num_bond = form_add.cleaned_data['num_bond']
            val_bond = form_add.cleaned_data['val_bond']
            state = form_add.cleaned_data['state']

            print(f"bidding: '{bidding_bond}'")
            print(f"idCard: '{id_card_bond}'")
            print(f"numBond: '{num_bond}'")
            print(f"value: '{val_bond}'")
            print(f"state: '{state}'")

            # Verificar si el número de bono ya existe
            if bindBonds.objects.filter(numBond=num_bond).exists():
                messages.error(request, "El número de bono ya está registrado. Por favor, ingresa otro número.")
                return render(request, "bid_bonds/add_bond.html", {
                    'form': form,
                    'form_add': form_add,
                })

            try:
                # Guardar el nuevo bono
                new_bond = bindBonds(
                    bidding=bidding_bond,
                    idCard=id_card_bond,
                    numBond=num_bond,
                    value=val_bond,
                    state=state
                )
                new_bond.save()

                messages.success(request, "Bono añadido correctamente.")
                return redirect('add_bond')  # Intentar redirigir a la lista de bonos

            except Exception as e:
                return render(request, "bid_bonds/add_bond.html", {
                    'form': form,
                    'form_add': form_add,
                })

        else:
            messages.error(request, "Por favor, corrija los errores en el formulario.")

    return render(request, "bid_bonds/add_bond.html", {
        'form': form,
        'form_add': form_add,
    })

@login_required
def delete_bond(request, bond_id):
    if bindBonds.objects.filter(id=bond_id).delete():
        messages.success(request, "Bono eliminado.")
    return redirect ('bond_management')

@login_required
def truncate_list(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute('TRUNCATE TABLE public."bindBonds" RESTART IDENTITY;')
        messages.success(request, 'Base limpiada correctamente')
    except Exception as e:
        messages.error(request, f"Error al limpiar la base {e}")
    return redirect('upload_bonds')