import re
from datetime import date
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.models import User
import base64


def convert_date(fecha):

    pattern = r'(\d+) de (\w+) de (\d{4})'
    match = re.match(pattern, fecha)

    if match:
        day = int(match.group(1))
        month_str = match.group(2)
        year = int(match.group(3))

        month_mapping = {
            'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
            'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
            'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
        }
        
        month = month_mapping.get(month_str.lower())

        if month:
            date_convert = date(year, month, day)
            return date_convert
        else:
            raise ValueError("Mes no válido")
    else:
        raise ValueError("Formato de fecha no válido")

def obtain_password (user):
    try:
        user = User.objects.get(username=user)
        return user.last_name
    except User.DoesNotExist:
        return None

def obtain_email (user):
    try:
        user = User.objects.get(username=user)
        return user.email
    except User.DoesNotExist:
        return None



def generate_excel_invoices (invoices, request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas manuales"

    headers = ["Número de factura", "Fecha de factura", "Fecha de carga", "Archivo", "Tienda"]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        ws[f"{col_letter}1"] = header
    
    for row_num, invoice in enumerate(invoices, 2):
        print(f"Empieza a generar data")
        ws[f"A{row_num}"] = invoice.numInvoice
        
        date_invoice_formatted = invoice.dateInvoice.strftime('%Y-%m-%d')
        ws[f"B{row_num}"] = date_invoice_formatted

        created = str(invoice.created)
        ws[f"C{row_num}"] = created

        file_url = request.build_absolute_uri('/')[:-1] + invoice.file.url
        ws[f"D{row_num}"].hyperlink = file_url 
        ws[f"D{row_num}"].value = invoice.file.name.split('/')[-1] 

        ws[f"E{row_num}"] = invoice.user.username
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
    
    wb.save(response)
    
    return response


def generate_excel_bonuses (sepBond, request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Bonos de separado"

    headers = ["Número de bono", "Fecha de factura", "Fecha de carga", "Valor del bono", "Archivo", "Tienda"]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        ws[f"{col_letter}1"] = header
    
    for row_num, bond in enumerate(sepBond, 2):
        print(f"Empieza a generar data")
        ws[f"A{row_num}"] = bond.numBond
        
        date_invoice_formatted = bond.dateInvoice.strftime('%Y-%m-%d')
        ws[f"B{row_num}"] = date_invoice_formatted

        created = str(bond.created)
        ws[f"C{row_num}"] = created

        valBond = str(bond.valBond)
        ws[f"D{row_num}"] = valBond
        
        file_url = request.build_absolute_uri('/')[:-1] + bond.file.url
        ws[f"E{row_num}"].hyperlink = file_url 
        ws[f"E{row_num}"].value = bond.file.name.split('/')[-1] 

        ws[f"F{row_num}"] = bond.user.username

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sepbonuses.xlsx"'
    
    wb.save(response)
    
    return response



def generate_excel_payDiscounts (payDiscount, request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Descuentos por nómina"

    headers = ["Número de factura", "Fecha de factura", "Fecha de carga", "Archivo", "Tienda"]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        ws[f"{col_letter}1"] = header
    
    for row_num, payd in enumerate(payDiscount, 2):
        print(f"Empieza a generar data")
        ws[f"A{row_num}"] = payd.numInvoice
        
        date_invoice_formatted = payd.datePayDiscounts.strftime('%Y-%m-%d')
        ws[f"B{row_num}"] = date_invoice_formatted

        created = str(payd.created)
        ws[f"C{row_num}"] = created
        
        file_url = request.build_absolute_uri('/')[:-1] + payd.file.url
        ws[f"D{row_num}"].hyperlink = file_url 
        ws[f"D{row_num}"].value = payd.file.name.split('/')[-1] 

        ws[f"E{row_num}"] = payd.user.username

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payDiscount.xlsx"'
    
    wb.save(response)
    
    return response



def generate_excel_dataManagement (data, request):
    print("Entra a generar excel")
    wb = Workbook()
    ws = wb.active
    ws.title = "Manejo de datos personales"

    headers = ["Número de cédula", "Fecha de carga", "Tienda", "Archivo"]
    print("Entra a generar encabezados")
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        ws[f"{col_letter}1"] = header
    
    for row_num, datam in enumerate(data, 2):
        print(f"Entra a generar data")
        ws[f"A{row_num}"] = datam.idCard
        
        date_formatted = datam.created.strftime('%Y-%m-%d')
        ws[f"B{row_num}"] = date_formatted

        ws[f"C{row_num}"] = datam.user.username
        
        file_url = request.build_absolute_uri('/')[:-1] + datam.file.url
        ws[f"D{row_num}"].hyperlink = file_url 
        ws[f"D{row_num}"].value = datam.file.name.split('/')[-1] 

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dataManagement.xlsx"'
    
    wb.save(response)
    
    return response

def manInvoices_dir_path(instance, filename):
    print ("Entra aca")
    return f'{instance.user.username}/manual_invoices/{filename}'





